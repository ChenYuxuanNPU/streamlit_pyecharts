import os
import shutil
import sqlite3

from openpyxl import Workbook
from openpyxl.styles import Border, Side, Font
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

db_path = r"C:\Users\1012986131\Desktop\python\streamlit_pyecharts\database\educational_data.db"

title_0 = [
    "单位全称（按照单位公章）", "学校类型", "统一社会信用代码", "姓名", "身份证号码（文本格式）", "性别",
    "民族", "籍贯（精确到市）", "婚姻状况", "政治面貌", "入党（团）时间", "参加工作前学历",
    "参加工作前学位", "参加工作前毕业院校（学信网全称）", "参加工作前毕业院校代码（五位数，中职或高中及以下请填‘无’）",
    "参加工作前毕业时间", "参加工作前所学专业全称", "最高学历", "最高学位", "最高学历毕业院校学信网全称",
    "最高学历毕业时间", "最高学历所学专业全称",
    "参加工作时间", "首次进入白云教育时间", "入白云区事业编时间", "进入现单位任教时间",
    "任现行政职务级别", "任现行政职务级别的时间", "任教年级",
    "主教学科（只能选一门）", "兼教学科(可以填多门，逗号隔开)", "工作情况备注", "现持有最高职称", "职称证专业",
    "最高职称证书认定或评审通过时间", "现聘专业技术职称最高级别", "现聘专业技术职级", "现聘专业技术职级首次聘用时间",
    "普通话层次", "教师资格层次", "教师资格学科", "最高骨干教师级别", "四名工作室主持人名称（最高级别）",
    "四名工作室主持人确定时间（最高级别）", "曾获市级及以上综合荣誉（可填多个，逗号隔开）",
    "曾获市级以上人才项目称号（可填多个，逗号隔开）", "支教单位全称（只填最近一次,限白云区教育局派出）",
    "支教地域（只填最近一次）", "支教开始时间（只填最近一次）", "支教结束时间（只填最近一次）",
    "交流单位全称（只填2015年以后且最近一次交流的情况，不包括跟岗和集中办公）", "交流开始时间", "交流结束时间",
    "现常住地址", "手机号码", "教育网短号", "紧急联系人", "与联系人关系", "紧急联系人手机号码", "备注",
    "个人信息填写备注",
    "年龄（两位数字，文本型）", "区域", "主要任教学段", "是否音体美专任教师", "党内职务级别", "校内其他职务",
    "职称证发证时间（xxxx年xx月，文本型）", "现聘职级（高级教师/一二级教师，不是专业技术级别）首次聘用时间（xxxx年xx月，文本型）"
    , "2024年9月编制所在学段", "2024年9月在岗状态"
]

# list_temp = ["广州培英教育集团", "广州市六十五中教育集团", "广州大同教育集团", "广州市白云中学教育集团",
#              "广州空港实验教育集团", "广外实验中学教育集团", "广州白云广附教育集团",
#              "广州白云六中教育集团", "广州云雅教育集团", "广州市白云区白云实验教育集团",
#              "广东第二师范学院实验教育集团", "广东技术师范大学实验教育集团", "三元里小学教育集团",
#              "广园小学教育集团", "景泰小学教育集团", "京溪小学教育集团", "华师附中实验小学教育集团",
#              "金沙第二小学教育集团", "广州民航幼儿园教育集团", "无", ]

list_b = ["十二年一贯制", "完全中学", "高中", "中职", "九年一贯制", "初中", "小学", "幼儿园", "教学支撑单位"]
list_f = ["男", "女"]
list_i = ["未婚", "已婚", "离异", "丧偶", "复婚", "再婚"]
list_j = ["群众", "中共党员", "共青团员", "民进会员", "民盟盟员", "致公党员", "中共预备党员", "民革会员",
          "九三学社", "农工党员", "无党派人士", "其他"]
list_l = ["博士研究生", "硕士研究生", "本科", "专科", "中专（非师范）", "中师", "高中", "初中"]
list_m = ["博士学位", "硕士学位", "学士学位", "无"]
list_r = ["博士研究生", "硕士研究生", "本科", "专科", "中专（非师范）", "中师", "高中", "初中"]
list_s = ["博士学位", "硕士学位", "学士学位", "无"]
list_aa = ["党组织书记", "党组织书记兼校长", "正校级", "副校级", "中层正职", "中层副职", "少先队大队辅导员",
           "少先队副大队辅导员", "工会主席", "工会副主席", "团委书记", "团委副书记", "无"]
list_ad = ["语文", "数学", "英语", "思想政治", "历史", "地理", "物理", "化学", "生物", "体育", "音乐", "美术",
           "书法", "舞蹈", "科学", "信息技术", "通用技术", "劳动", "综合实践", "心理健康", "人工智能", "汽修",
           "烹饪", "幼儿教育", "电子商务", "特殊教育", "校本课程", "其他", "无"]
list_ag = ["正高级教师", "高级教师", "一级教师", "二级教师", "三级教师", "高级职称（非中小学系列）",
           "中级职称（非中小学系列）", "初级职称（非中小学系列）", "未取得职称"]
list_aj = ["正高级教师", "高级教师", "一级教师", "二级教师", "三级教师", "高级职称（非中小学系列）",
           "中级职称（非中小学系列）", "初级职称（非中小学系列）", "试用期未聘", "未取得职称"]
list_ak = ["专业技术3级", "专业技术4级", "专业技术5级", "专业技术6级", "专业技术6级（暂聘）", "专业技术7级",
           "专业技术7级（暂聘）", "专业技术8级", "专业技术8级（暂聘）", "专业技术9级", "专业技术9级（暂聘）",
           "专业技术10级", "专业技术10级（暂聘）", "专业技术11级", "专业技术11级（暂聘）", "专业技术12级",
           "专业技术12级（暂聘）", "专业技术13级", "专业技术13级（暂聘）", "试用期（未定级）", "管理7级", "管理8级",
           "管理9级", "普通工", "中级工", "高级工", ]
list_am = ["一级甲等", "一级乙等", "二级甲等", "二级乙等", "三级甲等", "三级乙等", "无"]
list_an = ["幼儿园", "小学", "初级中学", "高级中学", "中职专业课", "中职实习指导教师", "高等学校", "无"]
list_ap = ["广东省骨干教师", "广州市骨干教师", "白云区骨干教师", "其他", "无"]
list_av = ["片内", "区内", "外市", "外省", "无"]
list_bk = ["直管", "永平", "石井", "新市", "江高", "人和", "太和", "钟落潭"]
list_bl = ["高中", "初中", "小学", "中职", "幼儿园", "其他"]
list_bm = ["是", "否"]
list_bn = ["党总支（党委）书记", "党总支（党委）副书记", "无"]
list_br = ["高中", "初中", "小学", "幼儿园", "职中", "其他"]
list_bs = ["在职在岗（本校任教）", "区内其他单位跟岗或借调（不在一线教学）", "区内交流（在一线教学）", "外市支教",
           "长期事假（7天以上）", "产假（不在岗）", "公假（丧假、婚假等）", "长期病休（不在岗）", "其他"]


def write_list_to_txt(route: str):
    with open(file=route, mode='w', encoding='UTF-8') as f:
        for i in (["B列"] + list_b):
            f.write(i + "\n")
        f.write("\n")

        for i in (["F列"] + list_f):
            f.write(i + "\n")
        f.write("\n")

        for i in (["I列"] + list_i):
            f.write(i + "\n")
        f.write("\n")

        for i in (["J列"] + list_j):
            f.write(i + "\n")
        f.write("\n")

        for i in (["L列"] + list_l):
            f.write(i + "\n")
        f.write("\n")

        for i in (["M列"] + list_m):
            f.write(i + "\n")
        f.write("\n")

        for i in (["R列"] + list_r):
            f.write(i + "\n")
        f.write("\n")

        for i in (["S列"] + list_s):
            f.write(i + "\n")
        f.write("\n")

        for i in (["AA列"] + list_aa):
            f.write(i + "\n")
        f.write("\n")

        for i in (["AD列"] + list_ad):
            f.write(i + "\n")
        f.write("\n")

        for i in (["AG列"] + list_ag):
            f.write(i + "\n")
        f.write("\n")

        for i in (["AJ列"] + list_aj):
            f.write(i + "\n")
        f.write("\n")

        for i in (["AK列"] + list_ak):
            f.write(i + "\n")
        f.write("\n")

        for i in (["AM列"] + list_am):
            f.write(i + "\n")
        f.write("\n")

        for i in (["AN列"] + list_an):
            f.write(i + "\n")
        f.write("\n")

        for i in (["AP列"] + list_ap):
            f.write(i + "\n")
        f.write("\n")

        for i in (["AV列"] + list_av):
            f.write(i + "\n")
        f.write("\n")

        for i in (["BK列"] + list_bk):
            f.write(i + "\n")
        f.write("\n")

        for i in (["BL列"] + list_bl):
            f.write(i + "\n")
        f.write("\n")

        for i in (["BM列"] + list_bm):
            f.write(i + "\n")
        f.write("\n")

        for i in (["BN列"] + list_bn):
            f.write(i + "\n")
        f.write("\n")

        for i in (["BR列"] + list_br):
            f.write(i + "\n")
        f.write("\n")

        for i in (["BS列"] + list_bs):
            f.write(i + "\n")
        f.write("\n")


def write_name_list_to_txt(route: str, name_list: list):
    with open(file=route, mode='w', encoding='UTF-8') as f:
        for name in name_list:
            f.write(name + "\n")


def ensure_folders_exist_or_clear(kind: str, area_name1="", school_name1=""):
    folder_names = ["直管", "永平", "江高", "石井", "新市", "人和", "太和", "钟落潭"]

    if school_name1 == "":
        base_dir = fr"C:\Users\1012986131\Desktop\python\streamlit_pyecharts\update_data_source\output\{kind}"

        """  
        确保在base_dir路径下存在列表中的每个文件夹，如果不存在则创建，如果存在则清空（删除并重建）。  

        :param base_dir: 文件夹的基路径  
        :param folder_names: 列表，包含要检查/创建的文件夹名称  
        """
        for folder_name in folder_names:
            full_path = os.path.join(base_dir, folder_name)

            # 检查文件夹是否存在
            if os.path.exists(full_path):
                # 如果存在，则删除并重新创建
                shutil.rmtree(full_path)  # 删除文件夹及其所有内容
                os.makedirs(full_path)  # 重新创建文件夹
            else:
                # 如果不存在，则创建文件夹
                os.makedirs(full_path)

    else:
        base_dir = fr"C:\Users\1012986131\Desktop\python\streamlit_pyecharts\update_data_source\output\{kind}\{area_name1}"
        full_path = os.path.join(base_dir, school_name1)

        # 检查文件夹是否存在
        if os.path.exists(full_path):
            # 如果存在，则删除并重新创建
            shutil.rmtree(full_path)  # 删除文件夹及其所有内容
            os.makedirs(full_path)  # 重新创建文件夹
        else:
            # 如果不存在，则创建文件夹
            os.makedirs(full_path)


def output_excel_0(title: list, data: list, file_name: str, area_name: str):
    def restrict(options: list, chara):

        # 创建一个数据验证对象，设置为列表类型，并设置允许的选项
        dv = DataValidation(type="list", formula1='"' + ','.join(map(str, options)) + '"')

        # 将数据验证添加到特定单元格（例如A1）
        for row in range(2, 500):  # 假设我们设置到第10行
            cell_ref = f'{chara}{row}'  # 构建单元格引用，如A1, A2, ..., A10
            # 添加到数据验证，然后将其应用到特定的单元格
            ws.add_data_validation(dv)
            dv.add(ws[cell_ref])

    def auto_adjust_column_width(worksheet, column):
        column_letter = get_column_letter(column)
        length = len(str(worksheet[column_letter][0].value))
        adjusted_width = (length + 8) * 1.5
        worksheet.column_dimensions[column_letter].width = adjusted_width

    def set_all_cell_borders(ws, border_style="thin", border_color="000000"):
        """
        为工作表中的所有单元格设置边框。

        :param ws: openpyxl.worksheet.worksheet.Worksheet 对象
        :param border_style: 边框的样式，如 "thin", "medium", "thick" 等
        :param border_color: 边框的颜色，格式为 "RRGGBB"
        """
        thin = Side(border_style=border_style, color=border_color)
        border = Border(top=thin, left=thin, right=thin, bottom=thin)

        # 遍历工作表中的每一行和每一列
        for row in ws.iter_rows():
            for cell in row:
                cell.border = border

    # 加入下拉选择框
    def add_restrict():
        restrict(options=list_b, chara="B")
        restrict(options=list_f, chara="F")
        restrict(options=list_i, chara="I")
        restrict(options=list_j, chara="J")
        restrict(options=list_l, chara="L")
        restrict(options=list_m, chara="M")
        restrict(options=list_r, chara="R")
        restrict(options=list_s, chara="S")
        restrict(options=list_aa, chara="AA")
        restrict(options=list_ad, chara="AD")
        restrict(options=list_ag, chara="AG")
        restrict(options=list_aj, chara="AJ")
        restrict(options=list_ak, chara="AK")
        restrict(options=list_am, chara="AM")
        restrict(options=list_an, chara="AN")
        restrict(options=list_ap, chara="AP")
        restrict(options=list_av, chara="AV")
        restrict(options=list_bk, chara="BK")
        restrict(options=list_bl, chara="BL")
        restrict(options=list_bm, chara="BM")
        restrict(options=list_bn, chara="BN")
        restrict(options=list_br, chara="BR")
        restrict(options=list_bs, chara="BS")

    # 加入背景色
    def add_bg_colour(ws):
        # 创建一个填充样式
        fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")

        # 应用填充样式到列A的所有单元格
        for cell in ws['B']:
            cell.fill = fill
        for cell in ws['I']:
            cell.fill = fill
        for cell in ws['J']:
            cell.fill = fill
        for cell in ws['L']:
            cell.fill = fill
        for cell in ws['R']:
            cell.fill = fill
        for cell in ws['AA']:
            cell.fill = fill
        for cell in ws['AK']:
            cell.fill = fill
        for cell in ws['AD']:
            cell.fill = fill

        # 创建一个填充样式
        fill = PatternFill(start_color="FFB6C1", end_color="FFB6C1", fill_type="solid")

        for cell in ws['BN']:
            cell.fill = fill
        for cell in ws['BO']:
            cell.fill = fill
        for cell in ws['BP']:
            cell.fill = fill
        for cell in ws['BQ']:
            cell.fill = fill
        for cell in ws['BR']:
            cell.fill = fill
        for cell in ws['BS']:
            cell.fill = fill

        # 创建一个填充样式
        fill = PatternFill(start_color="FFFFAA", end_color="FFFFAA", fill_type="solid")
        for cell in ws['F']:
            cell.fill = fill
        for cell in ws['M']:
            cell.fill = fill
        for cell in ws['S']:
            cell.fill = fill
        for cell in ws['AG']:
            cell.fill = fill
        for cell in ws['AJ']:
            cell.fill = fill
        for cell in ws['AM']:
            cell.fill = fill
        for cell in ws['AN']:
            cell.fill = fill
        for cell in ws['AP']:
            cell.fill = fill
        for cell in ws['AV']:
            cell.fill = fill
        for cell in ws['BK']:
            cell.fill = fill
        for cell in ws['BL']:
            cell.fill = fill
        for cell in ws['BM']:
            cell.fill = fill

    def set_col_width(ws):
        for i in range(len(title_0) + 7):
            auto_adjust_column_width(ws, i + 1)

    def add_excel_func():

        # 这里是用来判断填写的内容是否符合下拉列表的
        def gen_list_check_sentence(words: list, location: str):
            return f"{'+'.join([f'COUNTIF({location}:{location},"{value}")' for value in words])}"

        # 这里是用来判断理论上相同的一列（如校名）的总数，相当于填写的信息量
        def count_same_info(location: str):
            return f"=if(COUNTIF({location}:{location},{location}2)=COUNTA({location}:{location})-1,1,0)"

        # 这里用来判断位数是否统一
        def check_data_length(location: str, length: int):
            return f'=IF(SUMPRODUCT(--(LEN({location}:{location})={length}))=COUNTA(A:A)-1,1,0)'

        # 这里检查日期是不是按照xxxx年xx月格式
        def check_str_date(location: str):
            return f'=IF(COUNTIF({location}:{location},"????年??月")+COUNTIF({location}:{location},"????年?月")+COUNTIF({location}:{location},"无")=COUNTA(A:A)-1,1,0)'

        # 检查带函数的单元格是否被修改为常数
        def check_formula(td_list: list):
            formula = "=IF(ISFORMULA("
            formula += td_list[0]

            for i in range(1, len(td_list)):
                formula += f")*ISFORMULA({td_list[i]}"

            formula += "),1,0)"

            return formula

        # 这一段是用来判断是否有空单元格的，原理是遍历每一列，检查每一列的非空单元格数，若所有列的最大值和最小值相同，则认为所有信息填写完成
        def check_blank():
            # 初始化公式的基础部分
            formula = "=IF(MAX("

            # 初始化计数器，用于列名的生成
            column_letter = 'A'

            # 这里的四个范围，0，2代表A-Z，取值应该是(1,27)；1，3代表AA-A?，取值根据实际修改
            for i in range(1, 27):  # BS是Excel中的第702列
                # 生成列名（A-Z）
                column_name = chr(ord(column_letter) + i - 1)
                # 拼接COUNTA函数，注意添加范围（例如A1:A500）
                if i == 1:
                    formula += f"COUNTA({column_name}:{column_name})"
                else:
                    formula += f",COUNTA({column_name}:{column_name})"

            for i in range(1, 27):  # BS是Excel中的第702列
                # 生成列名（AA-AZ）
                column_name = chr(ord(column_letter) + i - 1)

                formula += f",COUNTA(A{column_name}:A{column_name})"

            for i in range(1, 20):  # BS是Excel中的第702列
                # 生成列名（AA-AZ）
                column_name = chr(ord(column_letter) + i - 1)

                formula += f",COUNTA(B{column_name}:B{column_name})"

            formula += ")=MIN("

            # 循环遍历从A到BS的所有列
            for i in range(1, 24):  # BS是Excel中的第702列
                # 生成列名（A-Z）
                column_name = chr(ord(column_letter) + i - 1)
                # 拼接COUNTA函数，注意添加范围（例如A1:A500）
                if i == 1:
                    formula += f"COUNTA({column_name}:{column_name})"
                else:
                    formula += f",COUNTA({column_name}:{column_name})"

            for i in range(1, 24):  # BS是Excel中的第702列
                # 生成列名（AA-AZ）
                column_name = chr(ord(column_letter) + i - 1)

                formula += f",COUNTA(A{column_name}:A{column_name})"

            for i in range(1, 20):  # BS是Excel中的第702列
                # 生成列名（AA-AZ）
                column_name = chr(ord(column_letter) + i - 1)

                formula += f",COUNTA(B{column_name}:B{column_name})"

            formula += "),1,0)"

            return formula

        # 为检查的字体添加红色
        font = Font(color="FF0000")

        # 填写完成的判断
        ws['BU1'] = '检查无误'
        ws[
            'BU2'] = "=IF(AND(BU5=1,BW2=1,BW5=1,BW8=1,BX2=1,BX5=1,BX8=1,BX11=1,BX14=1,BX17=1,BX20=1,BX23=1,BX26=1,BX29=1,BX32=1,BX35=1,BX38=1,BX41=1,BX44=1,BX47=1,BX50=1,BX53=1,BX56=1,BX59=1,BX62=1,BX65=1,BX68=1,BY2=1,BY5=1,BY8=1,BY11=1,BY14=1,BY17=1,BY20=1,BY23=1,BY26=1,BY29=1,BY32=1,BY35=1,BY38=1,BY41=1,BY44=1,BY47=1,BZ2=1,BZ5=1,BZ8=1),1,0)"
        ws['BU2'].font = font

        ws['BU4'] = '是否已填写完成'
        ws['BU5'] = check_blank()
        ws['BU5'].font = font

        ws['BU7'] = '公式是否完整'
        ws['BU8'] = check_formula(
            td_list=["BW2", "BW5", "BW8", "BX2", "BX5", "BX8", "BX11", "BX14", "BX17", "BX20", "BX23", "BX26",
                     "BX29", "BX32", "BX35", "BX38", "BX41", "BX44", "BX47", "BX50", "BX53", "BX56", "BX59", "BX62",
                     "BX65", "BX68", "BY2", "BY5", "BY8", "BY11", "BY14", "BY17", "BY20", "BY23", "BY26", "BY29",
                     "BY32", "BY35", "BY38", "BY41", "BY44", "BY47", "BZ2", "BZ5", "BZ8", "BU2", "BU5"])
        ws['BU8'].font = font

        # print(gen_list_check_sentence(words=["博士学位", "硕士学位", "学士学位", "无"],location="Y"))
        # 首先判断前三列学校信息会不会有不同的
        ws['BW1'] = 'A列是否相同'
        ws['BW2'] = count_same_info(location="A")
        ws['BW2'].font = font
        ws['BW4'] = 'B列是否相同'
        ws['BW5'] = count_same_info(location="B")
        ws['BW5'].font = font
        ws['BW7'] = 'C列是否相同'
        ws['BW8'] = count_same_info(location="C")
        ws['BW8'].font = font
        ws['BW10'] = 'BK列是否相同'
        ws['BW11'] = count_same_info(location="BK")
        ws['BW11'].font = font

        # 然后判断下拉列表的内容是否被更改
        ws['BX1'] = 'B列下拉列表'
        ws['BX2'] = f'=if(({gen_list_check_sentence(words=list_b, location="B")})=COUNTA(A:A)-1,1,0)'
        ws['BX2'].font = font

        ws['BX4'] = 'F列下拉列表'
        ws['BX5'] = f'=if(({gen_list_check_sentence(words=list_f, location="F")})=COUNTA(A:A)-1,1,0)'
        ws['BX5'].font = font

        ws['BX7'] = 'I列下拉列表'
        ws['BX8'] = f'=if(({gen_list_check_sentence(words=list_i, location="I")})=COUNTA(A:A)-1,1,0)'
        ws['BX8'].font = font

        ws['BX10'] = 'J列下拉列表'
        ws['BX11'] = f'=if(({gen_list_check_sentence(words=list_j, location="J")})=COUNTA(A:A)-1,1,0)'
        ws['BX11'].font = font

        ws['BX13'] = 'L列下拉列表'
        ws['BX14'] = f'=if(({gen_list_check_sentence(words=list_l, location="L")})=COUNTA(A:A)-1,1,0)'
        ws['BX14'].font = font

        ws['BX16'] = 'M列下拉列表'
        ws['BX17'] = f'=if(({gen_list_check_sentence(words=list_m, location="M")})=COUNTA(A:A)-1,1,0)'
        ws['BX17'].font = font

        ws['BX19'] = 'R列下拉列表'
        ws['BX20'] = f'=if(({gen_list_check_sentence(words=list_r, location="R")})=COUNTA(A:A)-1,1,0)'
        ws['BX20'].font = font

        ws['BX22'] = 'S列下拉列表'
        ws['BX23'] = f'=if(({gen_list_check_sentence(words=list_s, location="S")})=COUNTA(A:A)-1,1,0)'
        ws['BX23'].font = font

        ws['BX25'] = 'AA列下拉列表'
        ws['BX26'] = f'=if(({gen_list_check_sentence(words=list_aa, location="AA")})=COUNTA(A:A)-1,1,0)'
        ws['BX26'].font = font

        ws['BX28'] = 'AD列下拉列表'
        ws['BX29'] = f'=if(({gen_list_check_sentence(words=list_ad, location="AD")})=COUNTA(A:A)-1,1,0)'
        ws['BX29'].font = font

        ws['BX31'] = 'AG列下拉列表'
        ws['BX32'] = f'=if(({gen_list_check_sentence(words=list_ag, location="AG")})=COUNTA(A:A)-1,1,0)'
        ws['BX32'].font = font

        ws['BX34'] = 'AJ列下拉列表'
        ws['BX35'] = f'=if(({gen_list_check_sentence(words=list_aj, location="AJ")})=COUNTA(A:A)-1,1,0)'
        ws['BX35'].font = font

        ws['BX37'] = 'AK列下拉列表'
        ws['BX38'] = f'=if(({gen_list_check_sentence(words=list_ak, location="AK")})=COUNTA(A:A)-1,1,0)'
        ws['BX38'].font = font

        ws['BX40'] = 'AM列下拉列表'
        ws['BX41'] = f'=if(({gen_list_check_sentence(words=list_am, location="AM")})=COUNTA(A:A)-1,1,0)'
        ws['BX41'].font = font

        ws['BX43'] = 'AN列下拉列表'
        ws['BX44'] = f'=if(({gen_list_check_sentence(words=list_an, location="AN")})=COUNTA(A:A)-1,1,0)'
        ws['BX44'].font = font

        ws['BX46'] = 'AP列下拉列表'
        ws['BX47'] = f'=if(({gen_list_check_sentence(words=list_ap, location="AP")})=COUNTA(A:A)-1,1,0)'
        ws['BX47'].font = font

        ws['BX49'] = 'AV列下拉列表'
        ws['BX50'] = f'=if(({gen_list_check_sentence(words=list_av, location="AV")})=COUNTA(A:A)-1,1,0)'
        ws['BX50'].font = font

        ws['BX52'] = 'BK列下拉列表'
        ws['BX53'] = f'=if(({gen_list_check_sentence(words=list_bk, location="BK")})=COUNTA(A:A)-1,1,0)'
        ws['BX53'].font = font

        ws['BX55'] = 'BL列下拉列表'
        ws['BX56'] = f'=if(({gen_list_check_sentence(words=list_bl, location="BL")})=COUNTA(A:A)-1,1,0)'
        ws['BX56'].font = font

        ws['BX58'] = 'BM列下拉列表'
        ws['BX59'] = f'=if(({gen_list_check_sentence(words=list_bm, location="BM")})=COUNTA(A:A)-1,1,0)'
        ws['BX59'].font = font

        ws['BX61'] = 'BN列下拉列表'
        ws['BX62'] = f'=if(({gen_list_check_sentence(words=list_bn, location="BN")})=COUNTA(A:A)-1,1,0)'
        ws['BX62'].font = font

        ws['BX64'] = 'BR列下拉列表'
        ws['BX65'] = f'=if(({gen_list_check_sentence(words=list_br, location="BR")})=COUNTA(A:A)-1,1,0)'
        ws['BX65'].font = font

        ws['BX67'] = 'BS列下拉列表'
        ws['BX68'] = f'=if(({gen_list_check_sentence(words=list_bs, location="BS")})=COUNTA(A:A)-1,1,0)'
        ws['BX68'].font = font

        # 这里判断年月格式
        ws['BY1'] = 'K列日期格式'
        ws['BY2'] = check_str_date(location='K')
        ws['BY2'].font = font

        ws['BY4'] = 'P列日期格式'
        ws['BY5'] = check_str_date(location='P')
        ws['BY5'].font = font

        ws['BY7'] = 'U列日期格式'
        ws['BY8'] = check_str_date(location='U')
        ws['BY8'].font = font

        ws['BY10'] = 'W列日期格式'
        ws['BY11'] = check_str_date(location='W')
        ws['BY11'].font = font

        ws['BY13'] = 'X列日期格式'
        ws['BY14'] = check_str_date(location='X')
        ws['BY14'].font = font

        ws['BY16'] = 'Y列日期格式'
        ws['BY17'] = check_str_date(location='Y')
        ws['BY17'].font = font

        ws['BY19'] = 'Z列日期格式'
        ws['BY20'] = check_str_date(location='Z')
        ws['BY20'].font = font

        ws['BY22'] = 'AB列日期格式'
        ws['BY23'] = check_str_date(location='AB')
        ws['BY23'].font = font

        ws['BY25'] = 'AI列日期格式'
        ws['BY26'] = check_str_date(location='AI')
        ws['BY26'].font = font

        ws['BY28'] = 'AL列日期格式'
        ws['BY29'] = check_str_date(location='AL')
        ws['BY29'].font = font

        ws['BY31'] = 'AW列日期格式'
        ws['BY32'] = check_str_date(location='AW')
        ws['BY32'].font = font

        ws['BY34'] = 'AX列日期格式'
        ws['BY35'] = check_str_date(location='AX')
        ws['BY35'].font = font

        ws['BY37'] = 'AZ列日期格式'
        ws['BY38'] = check_str_date(location='AZ')
        ws['BY38'].font = font

        ws['BY40'] = 'BA列日期格式'
        ws['BY41'] = check_str_date(location='BA')
        ws['BY41'].font = font

        ws['BY43'] = 'BP列日期格式'
        ws['BY44'] = check_str_date(location='BP')
        ws['BY44'].font = font

        ws['BY46'] = 'BQ列日期格式'
        ws['BY47'] = check_str_date(location='BQ')
        ws['BY47'].font = font

        # 这里判断身份证位数
        ws['BZ1'] = 'E列位数'
        ws['BZ2'] = check_data_length(location='E', length=18)
        ws['BZ2'].font = font
        # 手机号位数
        ws['BZ4'] = 'BC列位数'
        ws['BZ5'] = check_data_length(location='BC', length=11)
        ws['BZ5'].font = font

        ws['BZ7'] = 'BG列位数'
        ws['BZ8'] = check_data_length(location='BG', length=11)
        ws['BZ8'].font = font

    # 这里是函数第一句
    wb = Workbook()
    ws = wb.active

    ws.append(title)

    area = data[0][-3]

    name_list = []
    for data_row in data:
        name_list.append(data_row[3])

    for data_row in data:
        ws.append(list(map(str, data_row)))

    add_restrict()

    add_bg_colour(ws=ws)

    add_excel_func()

    set_col_width(ws=ws)

    set_all_cell_borders(ws)

    ensure_folders_exist_or_clear(kind="在编", school_name1=file_name, area_name1=area_name)

    write_list_to_txt(
        route=fr"C:\Users\1012986131\Desktop\python\streamlit_pyecharts\update_data_source\output\在编\{area}\{file_name}\下拉列表内容.txt")

    write_name_list_to_txt(
        route=fr"C:\Users\1012986131\Desktop\python\streamlit_pyecharts\update_data_source\output\在编\{area}\{file_name}\已有人员名单.txt",
        name_list=name_list)

    wb.save(
        fr"C:\Users\1012986131\Desktop\python\streamlit_pyecharts\update_data_source\output\在编\{area}\{file_name}\{file_name}.xlsx")
    # print(file_name)


def del_tuple_in_list(data: list) -> list:
    # 删除第七项的出生年月
    return [[[x for i, x in enumerate(sublist) if (i != 6)] for sublist in mid_list] for mid_list in data]


if __name__ == '__main__':

    conn = sqlite3.connect(
        db_path
    )
    c = conn.cursor()

    # 搜集在编教师信息
    try:
        c.execute("select * from teacher_data_0")
        result0 = c.fetchall()

    except Exception as e:
        print('\033[1;91m' + f"执行mysql语句时报错：{e}" + '\033[0m')

    finally:
        conn.commit()

    # 检查学校总数
    try:
        c.execute("select distinct school_name from teacher_data_0")
        result_check = c.fetchall()
        print(len(result_check))

    except Exception as e:
        print('\033[1;91m' + f"执行mysql语句时报错：{e}" + '\033[0m')

    finally:
        conn.commit()
    # print(f"编内数据共包含{len(del_tuple_in_list(result_check))}所学校")

    conn.close()

    # 汇总相同学校的在编教师
    result0_all = []
    check_set = set()

    for data in result0:

        if data[0] not in check_set:
            result0_all.append([list(data)])

        else:
            for item in result0_all:
                if item[0][0] == data[0]:
                    item.append(list(data))

        check_set.add(data[0])

    print(len(result0_all))

    ensure_folders_exist_or_clear(kind="在编")

    result0_all = del_tuple_in_list(data=result0_all)

    for school in result0_all:
        excel_name = school[0][0]
        area_name = school[0][62]

        output_excel_0(title=title_0, data=school, file_name=excel_name, area_name=area_name)

        print(excel_name)

        break
