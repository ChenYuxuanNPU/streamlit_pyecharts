import os
import shutil
import sqlite3

from openpyxl import Workbook
from openpyxl.styles import Border, Side, Font
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

db_path = r"C:\Users\1012986131\Desktop\python\streamlit_pyecharts\database\educational_data.db"

title_1 = [
    "单位全称（按照单位公章）", "学校类型", "统一社会信用代码", "姓名", "身份证号码（文本格式）", "性别",
    "民族", "籍贯（精确到市）", "婚姻状况", "政治面貌", "入党（团）时间", "参加工作前学历",
    "参加工作前学位", "参加工作前毕业院校（学信网全称）", "参加工作前毕业时间（文本格式xxxx年xx月，其他日期同）",
    "参加工作前所学专业全称", "最高学历",
    "最高学位", "最高学历毕业院校学信网全称",
    "最高学历毕业时间", "最高学历所学专业全称", "参加工作时间", "首次进入白云教育时间", "进入现单位任教时间",
    "任现行政职务级别（如正校级、副校级、中层正副职）", "任现行政职务级别的时间",
    "任教年级", "主教学科（只能选一门）", "现持有最高职称", "职称证专业",
    "最高职称证书认定或评审通过时间", "普通话层次", "教师资格层次", "教师资格学科",
    "最高骨干教师级别", "四名工作室主持人名称（最高级别）",
    "四名工作室主持人确定时间（最高级别）", "曾获市级及以上综合荣誉（可填多个，逗号隔开）",
    "曾获市级以上人才项目称号（可填多个，逗号隔开）",
    "现常住地址", "手机号码", "紧急联系人", "与联系人关系", "紧急联系人手机号码", "备注", "区域",
    "任教学段", "是否音体美专任教师", "年龄", "教师身份"
]

list_b = ["十二年一贯制", "完全中学", "高中", "中职", "九年一贯制", "初中", "小学", "幼儿园", "教学支撑单位"]
list_f = ["男", "女"]
list_i = ["未婚", "已婚", "离异", "丧偶", "复婚", "再婚"]
list_j = ["群众", "中共党员", "共青团员", "民进会员", "民盟盟员", "致公党员", "中共预备党员", "民革会员",
          "九三学社", "农工党员", "无党派人士", "其他"]
list_l = ["博士研究生", "硕士研究生", "本科", "专科", "中专（非师范）", "中师", "高中", "初中"]
list_m = ["博士学位", "硕士学位", "学士学位", "无"]
list_q = ["博士研究生", "硕士研究生", "本科", "专科", "中专（非师范）", "中师", "高中", "初中"]
list_r = ["博士学位", "硕士学位", "学士学位", "无"]
list_y = ["党组织书记", "党组织书记兼校长", "正校级", "副校级", "中层正职", "中层副职", "少先队大队辅导员",
          "少先队副大队辅导员", "工会主席", "工会副主席", "团委书记", "团委副书记", "无"]
list_ab = ["语文", "数学", "英语", "思想政治", "历史", "地理", "物理", "化学", "生物", "体育", "音乐", "美术",
           "书法", "舞蹈", "科学", "信息技术", "通用技术", "劳动", "综合实践", "心理健康", "人工智能", "汽修",
           "烹饪", "幼儿教育", "电子商务", "特殊教育", "校本课程", "其他", "无"]
list_ac = ["正高级教师", "高级教师", "一级教师", "二级教师", "三级教师", "高级职称（非中小学系列）",
           "中级职称（非中小学系列）", "初级职称（非中小学系列）", "未取得职称"]
list_af = ["一级甲等", "一级乙等", "二级甲等", "二级乙等", "三级甲等", "三级乙等", "无"]
list_ag = ["幼儿园", "小学", "初级中学", "高级中学", "中职专业课", "中职实习指导教师", "高等学校", "无"]
list_ai = ["广东省骨干教师", "广州市骨干教师", "白云区骨干教师", "其他", "无"]
list_at = ["直管", "永平", "石井", "新市", "江高", "人和", "太和", "钟落潭"]
list_au = ["高中", "初中", "小学", "中职", "幼儿园", "其他"]
list_av = ["是", "否"]
list_ax = ["政府雇员", "公办临聘教师", "民办学校教师"]


def write_list_to_txt(route: str):
    with open(file=route, mode='w', encoding='UTF-8') as f:
        for i in (["B列"] + list_b):
            f.write(i + "\n")
        f.write("\n")

        for i in (["F列"] + list_f):
            f.write(i + "\n")
        f.write("\n")

        for i in (["I列"] + list_i):
            f.write(i + "\n")
        f.write("\n")

        for i in (["J列"] + list_j):
            f.write(i + "\n")
        f.write("\n")

        for i in (["L列"] + list_l):
            f.write(i + "\n")
        f.write("\n")

        for i in (["M列"] + list_m):
            f.write(i + "\n")
        f.write("\n")

        for i in (["Q列"] + list_q):
            f.write(i + "\n")
        f.write("\n")

        for i in (["R列"] + list_r):
            f.write(i + "\n")
        f.write("\n")

        for i in (["Y列"] + list_y):
            f.write(i + "\n")
        f.write("\n")

        for i in (["AB列"] + list_ab):
            f.write(i + "\n")
        f.write("\n")

        for i in (["AC列"] + list_ac):
            f.write(i + "\n")
        f.write("\n")

        for i in (["AF列"] + list_af):
            f.write(i + "\n")
        f.write("\n")

        for i in (["AG列"] + list_ag):
            f.write(i + "\n")
        f.write("\n")

        for i in (["AI列"] + list_ai):
            f.write(i + "\n")
        f.write("\n")

        for i in (["AT列"] + list_at):
            f.write(i + "\n")
        f.write("\n")

        for i in (["AU列"] + list_au):
            f.write(i + "\n")
        f.write("\n")

        for i in (["AV列"] + list_av):
            f.write(i + "\n")
        f.write("\n")

        for i in (["AX列"] + list_ax):
            f.write(i + "\n")
        f.write("\n")


def write_name_list_to_txt(route: str, name_list: list):
    with open(file=route, mode='w', encoding='UTF-8') as f:
        for name in name_list:
            f.write(name + "\n")


def ensure_folders_exist_or_clear(kind: str, area_name1="", school_name1=""):
    folder_names = ["直管", "永平", "江高", "石井", "新市", "人和", "太和", "钟落潭"]

    if school_name1 == "":
        base_dir = fr"C:\Users\1012986131\Desktop\python\streamlit_pyecharts\update_data_source\output\{kind}"

        """  
        确保在base_dir路径下存在列表中的每个文件夹，如果不存在则创建，如果存在则清空（删除并重建）。  

        :param base_dir: 文件夹的基路径  
        :param folder_names: 列表，包含要检查/创建的文件夹名称  
        """
        for folder_name in folder_names:
            full_path = os.path.join(base_dir, folder_name)

            # 检查文件夹是否存在
            if os.path.exists(full_path):
                # 如果存在，则删除并重新创建
                shutil.rmtree(full_path)  # 删除文件夹及其所有内容
                os.makedirs(full_path)  # 重新创建文件夹
            else:
                # 如果不存在，则创建文件夹
                os.makedirs(full_path)

    else:
        base_dir = fr"C:\Users\1012986131\Desktop\python\streamlit_pyecharts\update_data_source\output\{kind}\{area_name1}"
        full_path = os.path.join(base_dir, school_name1)

        # 检查文件夹是否存在
        if os.path.exists(full_path):
            # 如果存在，则删除并重新创建
            shutil.rmtree(full_path)  # 删除文件夹及其所有内容
            os.makedirs(full_path)  # 重新创建文件夹
        else:
            # 如果不存在，则创建文件夹
            os.makedirs(full_path)


def del_tuple_in_list(data: list) -> list:
    # 删除第七项的出生年月
    return [[[x for i, x in enumerate(sublist)] for sublist in mid_list] for mid_list in data]


def output_excel_1(title: list, data: list, file_name: str, area_name: str):
    def restrict(options: list, chara):

        # 创建一个数据验证对象，设置为列表类型，并设置允许的选项
        dv = DataValidation(type="list", formula1='"' + ','.join(map(str, options)) + '"')

        # 将数据验证添加到特定单元格（例如A1）
        for row in range(2, 1000):  # 假设我们设置到第1000行
            cell_ref = f'{chara}{row}'  # 构建单元格引用，如A1, A2, ..., A10
            # 添加到数据验证，然后将其应用到特定的单元格
            ws.add_data_validation(dv)
            dv.add(ws[cell_ref])

    def auto_adjust_column_width(worksheet, column):
        column_letter = get_column_letter(column)
        length = len(str(worksheet[column_letter][0].value))
        adjusted_width = (length + 8) * 1.5
        worksheet.column_dimensions[column_letter].width = adjusted_width

    def set_all_cell_borders(ws, border_style="thin", border_color="000000"):
        """
        为工作表中的所有单元格设置边框。

        :param ws: openpyxl.worksheet.worksheet.Worksheet 对象
        :param border_style: 边框的样式，如 "thin", "medium", "thick" 等
        :param border_color: 边框的颜色，格式为 "RRGGAZ"
        """
        thin = Side(border_style=border_style, color=border_color)
        border = Border(top=thin, left=thin, right=thin, bottom=thin)

        # 遍历工作表中的每一行和每一列
        for row in ws.iter_rows():
            for cell in row:
                cell.border = border

    # 加入下拉选择框
    def add_restrict():
        restrict(options=list_b, chara="B")
        restrict(options=list_f, chara="F")
        restrict(options=list_i, chara="I")
        restrict(options=list_j, chara="J")
        restrict(options=list_l, chara="L")
        restrict(options=list_m, chara="M")
        restrict(options=list_q, chara="Q")
        restrict(options=list_r, chara="R")
        restrict(options=list_y, chara="Y")
        restrict(options=list_ab, chara="AB")
        restrict(options=list_ac, chara="AC")
        restrict(options=list_af, chara="AF")
        restrict(options=list_ag, chara="AG")
        restrict(options=list_ai, chara="AI")
        restrict(options=list_at, chara="AT")
        restrict(options=list_au, chara="AU")
        restrict(options=list_av, chara="AV")
        restrict(options=list_ax, chara="AX")

    # 加入背景色
    def add_bg_colour(ws):
        # 创建一个填充样式
        fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")

        # 应用填充样式到列A的所有单元格
        for cell in ws['B']:
            cell.fill = fill
        for cell in ws['F']:
            cell.fill = fill
        for cell in ws['I']:
            cell.fill = fill
        for cell in ws['J']:
            cell.fill = fill
        for cell in ws['L']:
            cell.fill = fill
        for cell in ws['M']:
            cell.fill = fill
        for cell in ws['Q']:
            cell.fill = fill
        for cell in ws['R']:
            cell.fill = fill
        for cell in ws['Y']:
            cell.fill = fill
        for cell in ws['AB']:
            cell.fill = fill
        for cell in ws['AC']:
            cell.fill = fill
        for cell in ws['AF']:
            cell.fill = fill
        for cell in ws['AG']:
            cell.fill = fill
        for cell in ws['AI']:
            cell.fill = fill
        for cell in ws['AT']:
            cell.fill = fill
        for cell in ws['AU']:
            cell.fill = fill
        for cell in ws['AV']:
            cell.fill = fill
        for cell in ws['AX']:
            cell.fill = fill

    def set_col_width(ws):
        for i in range(len(title_1) + 10):
            auto_adjust_column_width(ws, i + 1)

    def add_excel_func():

        # 这里是用来判断填写的内容是否符合下拉列表的
        def gen_list_check_sentence(words: list, location: str):
            return f"{'+'.join([f'COUNTIF({location}:{location},"{value}")' for value in words])}"

        # 这里是用来判断理论上相同的一列（如校名）的总数，相当于填写的信息量
        def count_same_info(location: str):
            return f"=if(COUNTIF({location}:{location},{location}2)=COUNTA({location}:{location})-1,1,0)"

        # 这里用来判断位数是否统一
        def check_data_length(location: str, length: int):
            return f'=IF(SUMPRODUCT(--(LEN({location}:{location})={length}))=COUNTA(A:A)-1,1,0)'

        # 这里检查日期是不是按照xxxx年xx月格式
        def check_str_date(location: str):
            return f'=IF(COUNTIF({location}:{location},"????年??月")+COUNTIF({location}:{location},"????年?月")+COUNTIF({location}:{location},"无")=COUNTA(A:A)-1,1,0)'

        # 检查带函数的单元格是否被修改为常数
        def check_formula(td_list: list):
            formula = "=IF(ISFORMULA("
            formula += td_list[0]

            for i in range(1, len(td_list)):
                formula += f")*ISFORMULA({td_list[i]}"

            formula += "),1,0)"

            return formula

        # 这一段是用来判断是否有空单元格的，原理是遍历每一列，检查每一列的非空单元格数，若所有列的最大值和最小值相同，则认为所有信息填写完成
        def check_blank():
            # 初始化公式的基础部分
            formula = "=IF(MAX("

            # 初始化计数器，用于列名的生成
            column_letter = 'A'

            # 循环遍历从A到BS的所有列
            for i in range(1, 27):  # BS是Excel中的第702列
                # 生成列名（A-Z）
                column_name = chr(ord(column_letter) + i - 1)
                # 拼接COUNTA函数，注意添加范围（例如A1:A500）
                if i == 1:
                    formula += f"COUNTA({column_name}:{column_name})"
                else:
                    formula += f",COUNTA({column_name}:{column_name})"

            for i in range(1, 25):  # BS是Excel中的第702列
                # 生成列名（AA-AZ）
                column_name = chr(ord(column_letter) + i - 1)

                formula += f",COUNTA(A{column_name}:A{column_name})"

            formula += ")=MIN("

            # 循环遍历从A到BS的所有列
            for i in range(1, 27):  # BS是Excel中的第702列
                # 生成列名（A-Z）
                column_name = chr(ord(column_letter) + i - 1)
                # 拼接COUNTA函数，注意添加范围（例如A1:A500）
                if i == 1:
                    formula += f"COUNTA({column_name}:{column_name})"
                else:
                    formula += f",COUNTA({column_name}:{column_name})"

            for i in range(1, 25):  # BS是Excel中的第702列
                # 生成列名（AA-AZ）
                column_name = chr(ord(column_letter) + i - 1)

                formula += f",COUNTA(A{column_name}:A{column_name})"

            formula += "),1,0)"

            return formula

        # 为检查的字体添加红色
        font = Font(color="FF0000")

        # 填写完成的判断
        ws['AZ1'] = '检查无误'
        ws[
            'AZ2'] = "=IF(AND(AZ5=1,AZ8=1,BB2=1,BB5=1,BB8=1,BB11=1,BD2=1,BD5=1,BD8=1,BD11=1,BD14=1,BD17=1,BD20=1,BD23=1,BD26=1,BD29=1,BD32=1,BD35=1,BD38=1,BD41=1,BD44=1,BD47=1,BF2=1,BF5=1,BF8=1,BF11=1,BF14=1,BF17=1,BF20=1,BF23=1,BF26=1,BH2=1,BH5=1,BH8=1),1,0)"
        ws['AZ2'].font = font

        ws['AZ4'] = '是否已填写完成'
        ws['AZ5'] = check_blank()
        ws['AZ5'].font = font

        ws['AZ7'] = '公式是否完整'
        ws['AZ8'] = check_formula(
            td_list=["AZ8", "BB2", "BB5", "BB8", "BB11", "BD2", "BD5", "BD8", "BD11", "BD14", "BD17", "BD20",
                     "BD23", "BD26", "BD29", "BD32", "BD35", "BD38", "BD41", "BD44", "BD47",
                     "BF2", "BF5", "BF8", "BF11", "BF14", "BF17", "BF20", "BF23", "BF26", "BH2", "BH5", "BH8"])
        ws['AZ8'].font = font

        ws['AZ10'] = '空单元格个数(需要调整AX后行序号)'
        ws['AZ11'] = "=COUNTBLANK(A2:AX100)"
        ws['AZ11'].font = font

        # print(gen_list_check_sentence(words=["博士学位", "硕士学位", "学士学位", "无"],location="Y"))
        # 首先判断前三列学校信息会不会有不同的
        ws['BB1'] = 'A列是否相同'
        ws['BB2'] = count_same_info(location="A")
        ws['BB2'].font = font
        ws['BB4'] = 'B列是否相同'
        ws['BB5'] = count_same_info(location="B")
        ws['BB5'].font = font
        ws['BB7'] = 'C列是否相同'
        ws['BB8'] = count_same_info(location="C")
        ws['BB8'].font = font
        ws['BB10'] = 'AT列是否相同'
        ws['BB11'] = count_same_info(location="AT")
        ws['BB11'].font = font

        # 然后判断下拉列表的内容是否被更改
        ws['BD1'] = 'B列下拉列表'
        ws['BD2'] = f'=if(({gen_list_check_sentence(words=list_b, location="B")})=COUNTA(A:A)-1,1,0)'
        ws['BD2'].font = font

        ws['BD4'] = 'F列下拉列表'
        ws['BD5'] = f'=if(({gen_list_check_sentence(words=list_f, location="F")})=COUNTA(A:A)-1,1,0)'
        ws['BD5'].font = font

        ws['BD7'] = 'I列下拉列表'
        ws['BD8'] = f'=if(({gen_list_check_sentence(words=list_i, location="I")})=COUNTA(A:A)-1,1,0)'
        ws['BD8'].font = font

        ws['BD10'] = 'J列下拉列表'
        ws['BD11'] = f'=if(({gen_list_check_sentence(words=list_j, location="J")})=COUNTA(A:A)-1,1,0)'
        ws['BD11'].font = font

        ws['BD13'] = 'L列下拉列表'
        ws['BD14'] = f'=if(({gen_list_check_sentence(words=list_l, location="L")})=COUNTA(A:A)-1,1,0)'
        ws['BD14'].font = font

        ws['BD16'] = 'M列下拉列表'
        ws['BD17'] = f'=if(({gen_list_check_sentence(words=list_m, location="M")})=COUNTA(A:A)-1,1,0)'
        ws['BD17'].font = font

        ws['BD19'] = 'Q列下拉列表'
        ws['BD20'] = f'=if(({gen_list_check_sentence(words=list_q, location="Q")})=COUNTA(A:A)-1,1,0)'
        ws['BD20'].font = font

        ws['BD22'] = 'R列下拉列表'
        ws['BD23'] = f'=if(({gen_list_check_sentence(words=list_r, location="R")})=COUNTA(A:A)-1,1,0)'
        ws['BD23'].font = font

        ws['BD25'] = 'Y列下拉列表'
        ws['BD26'] = f'=if(({gen_list_check_sentence(words=list_y, location="Y")})=COUNTA(A:A)-1,1,0)'
        ws['BD26'].font = font

        ws['BD28'] = 'AB列下拉列表'
        ws['BD29'] = f'=if(({gen_list_check_sentence(words=list_ab, location="AB")})=COUNTA(A:A)-1,1,0)'
        ws['BD29'].font = font

        ws['BD31'] = 'AC列下拉列表'
        ws['BD32'] = f'=if(({gen_list_check_sentence(words=list_ac, location="AC")})=COUNTA(A:A)-1,1,0)'
        ws['BD32'].font = font

        ws['BD34'] = 'AF列下拉列表'
        ws['BD35'] = f'=if(({gen_list_check_sentence(words=list_af, location="AF")})=COUNTA(A:A)-1,1,0)'
        ws['BD35'].font = font

        ws['BD37'] = 'AG列下拉列表'
        ws['BD38'] = f'=if(({gen_list_check_sentence(words=list_ag, location="AG")})=COUNTA(A:A)-1,1,0)'
        ws['BD38'].font = font

        ws['BD40'] = 'AI列下拉列表'
        ws['BD41'] = f'=if(({gen_list_check_sentence(words=list_ai, location="AI")})=COUNTA(A:A)-1,1,0)'
        ws['BD41'].font = font

        ws['BD43'] = 'AT列下拉列表'
        ws['BD44'] = f'=if(({gen_list_check_sentence(words=list_at, location="AT")})=COUNTA(A:A)-1,1,0)'
        ws['BD44'].font = font

        ws['BD46'] = 'AU列下拉列表'
        ws['BD47'] = f'=if(({gen_list_check_sentence(words=list_au, location="AU")})=COUNTA(A:A)-1,1,0)'
        ws['BD47'].font = font

        ws['BD46'] = 'AV列下拉列表'
        ws['BD47'] = f'=if(({gen_list_check_sentence(words=list_av, location="AV")})=COUNTA(A:A)-1,1,0)'
        ws['BD47'].font = font

        ws['BD46'] = 'AX列下拉列表'
        ws['BD47'] = f'=if(({gen_list_check_sentence(words=list_ax, location="AX")})=COUNTA(A:A)-1,1,0)'
        ws['BD47'].font = font

        # 这里判断年月格式
        ws['BF1'] = 'K列日期格式'
        ws['BF2'] = check_str_date(location='K')
        ws['BF2'].font = font

        ws['BF4'] = 'O列日期格式'
        ws['BF5'] = check_str_date(location='O')
        ws['BF5'].font = font

        ws['BF7'] = 'T列日期格式'
        ws['BF8'] = check_str_date(location='T')
        ws['BF8'].font = font

        ws['BF10'] = 'V列日期格式'
        ws['BF11'] = check_str_date(location='V')
        ws['BF11'].font = font

        ws['BF13'] = 'W列日期格式'
        ws['BF14'] = check_str_date(location='W')
        ws['BF14'].font = font

        ws['BF16'] = 'X列日期格式'
        ws['BF17'] = check_str_date(location='X')
        ws['BF17'].font = font

        ws['BF19'] = 'Z列日期格式'
        ws['BF20'] = check_str_date(location='Z')
        ws['BF20'].font = font

        ws['BF22'] = 'AE列日期格式'
        ws['BF23'] = check_str_date(location='AE')
        ws['BF23'].font = font

        ws['BF25'] = 'AK列日期格式'
        ws['BF26'] = check_str_date(location='AK')
        ws['BF26'].font = font

        # 这里判断身份证位数
        ws['BH1'] = 'E列位数'
        ws['BH2'] = check_data_length(location='E', length=18)
        ws['BH2'].font = font
        # 手机号位数
        ws['BH4'] = 'AO列位数'
        ws['BH5'] = check_data_length(location='AO', length=11)
        ws['BH5'].font = font

        ws['BH7'] = 'AR列位数'
        ws['BH8'] = check_data_length(location='AR', length=11)
        ws['BH8'].font = font

    # 这里是函数第一句
    wb = Workbook()
    ws = wb.active

    ws.append(title)

    area = data[0][-4]

    name_list = []
    for data_row in data:
        name_list.append(data_row[3])

    for data_row in data:
        ws.append(list(map(str, data_row)))

    add_restrict()

    add_bg_colour(ws=ws)

    add_excel_func()

    set_col_width(ws=ws)

    set_all_cell_borders(ws)

    ensure_folders_exist_or_clear(kind="编外", school_name1=file_name, area_name1=area_name)

    write_list_to_txt(
        route=fr"C:\Users\1012986131\Desktop\python\streamlit_pyecharts\update_data_source\output\编外\{area}\{file_name}\下拉列表内容.txt")

    write_name_list_to_txt(
        route=fr"C:\Users\1012986131\Desktop\python\streamlit_pyecharts\update_data_source\output\编外\{area}\{file_name}\已有人员名单.txt",
        name_list=name_list)

    wb.save(
        fr"C:\Users\1012986131\Desktop\python\streamlit_pyecharts\update_data_source\output\编外\{area}\{file_name}\{file_name}.xlsx")
    # print(file_name)


if __name__ == '__main__':

    conn = sqlite3.connect(
        db_path
    )
    c = conn.cursor()

    # 搜集在编教师信息
    try:
        c.execute("select * from teacher_data_1_2023")
        result1 = c.fetchall()

    except Exception as e:
        print('\033[1;91m' + f"执行mysql语句时报错：{e}" + '\033[0m')

    finally:
        conn.commit()

    # 检查学校总数
    try:
        c.execute('select distinct "校名" from teacher_data_1_2023')
        result_check = c.fetchall()

    except Exception as e:
        print('\033[1;91m' + f"执行mysql语句时报错：{e}" + '\033[0m')

    finally:
        conn.commit()
    print(f"编外数据共包含{len(del_tuple_in_list(result_check))}所学校")

    conn.close()

    # 汇总相同学校的在编教师
    result1_all = []
    check_set = set()

    for data in result1:

        if data[0] not in check_set:
            result1_all.append([list(data)])

        else:
            for item in result1_all:
                if item[0][0] == data[0]:
                    item.append(list(data))

        check_set.add(data[0])

    result1_all = del_tuple_in_list(data=result1_all)

    ensure_folders_exist_or_clear(kind="编外")

    for school in result1_all:
        excel_name = school[0][0]
        area_name = school[0][-4]

        output_excel_1(title=title_1, data=school, file_name=excel_name, area_name=area_name)

        print(excel_name)

        break