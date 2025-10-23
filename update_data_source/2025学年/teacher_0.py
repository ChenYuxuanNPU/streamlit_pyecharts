import json
import os
import shutil

from openpyxl import Workbook
from openpyxl.styles import Border, Side, Font
from openpyxl.styles import PatternFill
from openpyxl.styles import numbers
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.protection import SheetProtection

from func import read_xlsx_to_list, generate_subsets

# 下次用的时候，记得限制本科及以上不能不选或填学校

title_0 = [
    "单位全称（按照单位公章）", "学校类型", "统一社会信用代码", "姓名", "身份证号码（文本格式）", "性别",
    "民族", "籍贯（精确到市）", "婚姻状况", "政治面貌", "入党（团）时间", "参加工作前学历",
    "参加工作前学位", "参加工作前毕业院校（普通高校）", "参加工作前毕业院校（境外高校或其他院校）",
    "参加工作前毕业时间", "参加工作前所学专业全称", "最高学历", "最高学历对应学位", "最高学历毕业院校（普通高校）",
    "最高学历毕业院校（境外高校或其他院校）", "最高学历毕业时间", "最高学历所学专业全称",
    "参加工作时间", "首次进入白云教育时间", "入白云区事业编时间", "进入现单位任教时间",
    "行政职务", "任现行政职务级别的时间", "任教年级（从低到高填写，按照列表要求）",
    "主教学科（只能选一门）", "兼教学科(可以填多门，逗号隔开)", "现持有最高职称", "职称证专业（填学段学科）",
    "职称证发证时间", "最高职称证书认定或评审通过时间", "现聘专业技术职称",
    "现聘职级（高级教师/一二级教师，不是专业技术级别）首次聘用时间",
    "现聘专业技术岗位", "现聘专业技术岗位首次聘用时间",
    "普通话层次", "教师资格层次", "教师资格学科", "最高骨干教师级别", "四名工作室主持人名称（最高级别）",
    "四名工作室主持人确定时间（最高级别）", "曾获市级及以上综合荣誉（可填多个，逗号隔开）",
    "曾获市级以上人才项目称号（可填多个，逗号隔开）", "支教单位全称（只填最近一次,限白云区教育局派出）",
    "支教地域（只填最近一次）", "支教开始时间（只填最近一次）", "支教结束时间（只填最近一次）",
    "交流单位全称（只填2015年以后且最近一次交流的情况，不包括跟岗和集中办公）", "交流开始时间", "交流结束时间",
    "现常住地址", "手机号码", "教育网短号", "紧急联系人", "与联系人关系", "紧急联系人手机号码",
    "区域", "主要任教学段", "2025年9月编制所在学段", "2025年9月在岗状态", "是否音体美专任教师", "党内职务级别",
    "校内其他职务", "备注", "非学历教育最高学位（如同等学力申硕）", "非学历教育最高学位毕业院校",
    "非学历教育最高学位毕业院校（境外高校或其他院校）"
]
lst = {
    "list_b": ["十二年一贯制", "完全中学", "高中", "中职", "九年一贯制", "初中", "小学", "幼儿园", "特殊教育学校",
               "教学支撑单位"],
    "list_f": ["男", "女"],
    "list_g": ["汉族", "蒙古族", "回族", "藏族", "维吾尔族", "苗族", "彝族", "壮族", "布依族", "朝鲜族", "满族", "侗族",
               "瑶族", "白族", "土家族", "哈尼族", "哈萨克族", "傣族", "黎族", "傈僳族", "佤族", "畲族", "高山族",
               "拉祜族", "水族", "东乡族", "纳西族", "景颇族", "柯尔克孜族", "土族", "达斡尔族", "仫佬族", "羌族",
               "布朗族", "撒拉族", "毛南族", "仡佬族", "锡伯族", "阿昌族", "普米族", "塔吉克族", "怒族", "乌孜别克族",
               "俄罗斯族", "鄂温克族", "德昂族", "保安族", "裕固族", "京族", "塔塔尔族", "独龙族", "鄂伦春族", "赫哲族",
               "门巴族", "珞巴族", "基诺族"],
    "list_i": ["未婚", "已婚", "离异", "丧偶", "复婚", "再婚"],
    "list_j": ["群众", "中共党员", "共青团员", "民进会员", "民盟盟员", "致公党员", "中共预备党员", "民革会员",
               "九三学社", "农工党员", "无党派人士", "其他"],
    "list_l": ["博士研究生", "硕士研究生", "本科", "专科", "中专（非师范）", "中师", "高中", "初中"],
    "list_m": ["博士学位", "硕士学位", "学士学位", "无"],
    "list_n": ["其他", "无"] + [item[0] for item in read_xlsx_to_list(
        r"C:\Users\1012986131\Desktop\python\streamlit_pyecharts\参考资料\2025年6月院校名单.xlsx")[1:]],
    "list_r": ["博士研究生", "硕士研究生", "本科", "专科", "中专（非师范）", "中师", "高中", "初中"],
    "list_s": ["博士学位", "硕士学位", "学士学位", "无"],
    "list_t": ["其他", "无"] + [item[0] for item in read_xlsx_to_list(
        r"C:\Users\1012986131\Desktop\python\streamlit_pyecharts\参考资料\2025年6月院校名单.xlsx")[1:]],
    "list_ab": ["书记（正校级）", "副书记（副校级）", "正校级", "副校级", "中层正职", "中层副职", "少先队大队辅导员",
                "少先队副大队辅导员", "工会主席", "工会副主席", "团委书记", "团委副书记", "无"],
    "list_ad": generate_subsets(
        ["幼儿园", "一年级", "二年级", "三年级", "四年级", "五年级", "六年级", "七年级", "八年级",
         "九年级", "高一年级", "高二年级", "高三年级", "中职一年级", "中职二年级", "中职三年级"]) + ["无"],
    "list_ae": ["语文", "数学", "英语", "思想政治", "历史", "地理", "物理", "化学", "生物", "体育", "音乐", "美术",
                "书法", "舞蹈", "科学", "信息技术", "通用技术", "劳动", "综合实践", "心理健康", "人工智能", "汽修",
                "烹饪", "幼儿教育", "电子商务", "特殊教育", "校本课程", "其他", "无"],
    "list_ag": ["正高级教师", "高级教师", "一级教师", "二级教师", "三级教师", "高级职称（非中小学系列）",
                "中级职称（非中小学系列）", "初级职称（非中小学系列）", "未取得职称"],
    "list_ak": ["正高级教师", "高级教师", "一级教师", "二级教师", "三级教师", "高级职称（非中小学系列）",
                "中级职称（非中小学系列）", "初级职称（非中小学系列）", "试用期未聘", "未取得职称"],
    "list_am": ["专业技术三级", "专业技术四级", "专业技术五级", "专业技术六级", "专业技术六级（暂聘）", "专业技术七级",
                "专业技术七级（暂聘）", "专业技术八级", "专业技术八级（暂聘）", "专业技术九级", "专业技术九级（暂聘）",
                "专业技术十级", "专业技术十级（暂聘）", "专业技术十一级", "专业技术十一级（暂聘）", "专业技术十二级",
                "专业技术十二级（暂聘）", "专业技术十三级", "专业技术十三级（暂聘）", "试用期（未定级）", "管理七级",
                "管理八级", "管理九级", "普通工", "中级工", "高级工", ],
    "list_ao": ["一级甲等", "一级乙等", "二级甲等", "二级乙等", "三级甲等", "三级乙等", "无"],
    "list_ap": ["幼儿园", "小学", "初级中学", "高级中学", "中职专业课", "中职实习指导教师", "高等学校", "无"],
    "list_ar": ["广东省骨干教师", "广州市骨干教师", "白云区骨干教师", "其他", "无"],
    "list_ax": ["片内", "区内", "外市", "外省", "无"],
    "list_bj": ["直管", "永平", "石井", "新市", "江高", "人和", "太和", "钟落潭"],
    "list_bk": ["高中", "初中", "小学", "中职", "幼儿园", "其他"],
    "list_bl": ["高中", "初中", "小学", "幼儿园", "职中", "其他"],
    "list_bm": ["在职在岗（本校任教）", "区内其他单位跟岗或借调（不在一线教学）", "区内交流（在一线教学）", "外市支教",
                "长期事假（7天以上）", "产假（不在岗）", "公假（丧假、婚假等）", "长期病休（不在岗）", "其他"],
    "list_bn": ["是", "否"],
    "list_bo": ["党总支（党委）书记", "党总支（党委）副书记", "无"],
    "list_br": ["博士学位", "硕士学位", "学士学位", "无"],
    "list_bs": ["其他", "无"] + [item[0] for item in read_xlsx_to_list(
        r"C:\Users\1012986131\Desktop\python\streamlit_pyecharts\参考资料\2025年6月院校名单.xlsx")[1:]],
}


def letter_to_number(letter_str):
    """
    将字母字符串转换为数字
    规则：A=1, B=2, ..., Z=26, AA=27, AB=28, ..., ZZ=702, AAA=703, 等等
    """
    if not letter_str or not letter_str.isalpha():
        return None

    letter_str = letter_str.upper()
    total = 0

    for char in letter_str:
        # 将当前字符转换为数字 (A=1, B=2, ..., Z=26)
        char_value = ord(char) - ord('A') + 1
        # 将之前的结果乘以26（相当于进位），然后加上当前字符的值
        total = total * 26 + char_value

    return total


def write_list_to_txt(route: str):
    with open(file=route, mode='w', encoding='UTF-8') as f:
        for lst_temp in lst.keys():
            for i in ([f"{lst_temp.split("_")[-1].upper()}列"] + lst[lst_temp]):
                f.write(str(i) + "\n")
        f.write("\n")


def write_name_list_to_txt(route: str, name_list: list):
    with open(file=route, mode='w', encoding='UTF-8') as f:
        for name in name_list:
            f.write(name + "\n")


def ensure_folders_exist_or_clear(kind: str, area_name1="", school_name1=""):
    folder_names = ["直管", "永平", "江高", "石井", "新市", "人和", "太和", "钟落潭"]

    if school_name1 == "":
        base_dir = fr"C:\Users\1012986131\Desktop\python\streamlit_pyecharts\update_data_source\2025学年\output\{kind}"

        """  
        确保在base_dir路径下存在列表中的每个文件夹，如果不存在则创建，如果存在则清空（删除并重建）。  

        :param base_dir: 文件夹的基路径  
        :param folder_names: 列表，包含要检查/创建的文件夹名称  
        """
        for folder_name in folder_names:
            full_path = os.path.join(base_dir, folder_name)

            # 检查文件夹是否存在
            if os.path.exists(full_path):
                # 如果存在，则删除并重新创建
                shutil.rmtree(full_path)  # 删除文件夹及其所有内容
                os.makedirs(full_path)  # 重新创建文件夹
            else:
                # 如果不存在，则创建文件夹
                os.makedirs(full_path)

    else:
        base_dir = fr"C:\Users\1012986131\Desktop\python\streamlit_pyecharts\update_data_source\2025学年\output\{kind}\{area_name1}"
        full_path = os.path.join(base_dir, school_name1)

        # 检查文件夹是否存在
        if os.path.exists(full_path):
            # 如果存在，则删除并重新创建
            shutil.rmtree(full_path)  # 删除文件夹及其所有内容
            os.makedirs(full_path)  # 重新创建文件夹
        else:
            # 如果不存在，则创建文件夹
            os.makedirs(full_path)


# todo


def output_excel_0(title: list, data: list, file_name: str, area_name: str):
    def add_sheet_limit(sheet):
        protection = SheetProtection(
            sheet=True,  # 启用保护
            selectLockedCells=False,  # 禁止选择锁定单元格
            selectUnlockedCells=False,  # 禁止选择未锁定单元格
            formatCells=False,  # 禁止格式化单元格
            formatColumns=False,  # 禁止格式化列
            formatRows=False,  # 禁止格式化行
            insertColumns=False,  # 禁止插入列
            insertRows=False,  # 禁止插入行
            insertHyperlinks=False,  # 禁止插入超链接
            deleteColumns=False,  # 禁止删除列
            deleteRows=False,  # 禁止删除行
            sort=False,  # 禁止排序
            autoFilter=False,  # 禁止自动筛选
            pivotTables=False,  # 禁止数据透视表
            objects=False,  # 禁止操作对象
            scenarios=False  # 禁止方案
        )

        protection.password = 'byjyy'

        sheet.protection = protection

    def restrict(options: list, chara):

        for i, item in enumerate(options, 1):
            ws_list.cell(row=i, column=letter_to_number(chara)).value = item

        data_range = ("'{}'!" + f"${chara}:${chara}").format("下拉列表信息")
        dv = DataValidation(type="list", formula1="=" + data_range, showDropDown=False)

        for row in range(2, 500):
            dv.add(ws["{}{}".format(f"{chara}", row)])

        ws.add_data_validation(dv)

    def auto_adjust_column_width(worksheet, column):
        column_letter = get_column_letter(column)
        length = len(str(worksheet[column_letter][0].value))
        adjusted_width = (length + 8) * 1.45
        worksheet.column_dimensions[column_letter].width = adjusted_width

    def set_all_cell_borders(ws, border_style="thin", border_color="000000"):
        """
        为工作表中的所有单元格设置边框。

        :param ws: openpyxl.worksheet.worksheet.Worksheet 对象
        :param border_style: 边框的样式，如 "thin", "medium", "thick" 等
        :param border_color: 边框的颜色，格式为 "RRGGBB"
        """
        thin = Side(border_style=border_style, color=border_color)
        border = Border(top=thin, left=thin, right=thin, bottom=thin)

        # 遍历工作表中的每一行和每一列
        for row in ws.iter_rows():
            for cell in row:
                cell.border = border

    # 加入下拉选择框
    def add_restrict():
        for lst_temp in lst.keys():
            restrict(options=lst[lst_temp], chara=lst_temp.split("_")[-1].upper())

    # 加入背景色
    def add_bg_colour(ws):
        # 创建一个填充样式
        fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")

        for lst_temp in lst.keys():
            for cell in ws[lst_temp.split("_")[-1].upper()]:
                cell.fill = fill

    def set_col_width(ws):
        for i in range(len(title_0)):
            auto_adjust_column_width(ws, i + 1)

    def add_excel_func():

        cell_list = []

        # 这里是用来判断填写的内容是否符合下拉列表的
        def gen_list_check_sentence(location: str):
            return f"=if(SUMPRODUCT(COUNTIF(下拉列表信息!{location.upper()}:{location.upper()},数据表!{location.upper()}:{location.upper()}))=COUNTA(数据表!{location}:{location})-1,1,0)"

        # 这里是用来判断理论上相同的一列（如校名）的总数，相当于填写的信息量
        def count_same_info(location: str):
            return f"=if(COUNTIF(数据表!{location}:{location},数据表!{location}2)=COUNTA(数据表!{location}:{location})-1,1,0)"

        # 这里用来判断位数是否统一
        def check_data_length(location: str, length: int):
            return f'=IF(SUMPRODUCT(--(LEN(数据表!{location}:{location})={length}))=COUNTA(数据表!A:A)-1,1,0)'

        # 这里检查日期是不是按照xxxx年xx月格式
        def check_str_date(location: str):
            return f'=IF(COUNTIF(数据表!{location}:{location},"????年??月")+COUNTIF(数据表!{location}:{location},"????年?月")+COUNTIF(数据表!{location}:{location},"无")=COUNTA(数据表!A:A)-1,1,0)'

        # 检查带函数的单元格是否被修改为常数
        def check_formula(td_list: list):
            formula = "=IF(ISFORMULA("
            formula += td_list[0]

            for i in range(1, len(td_list)):
                formula += f")*ISFORMULA({td_list[i]}"

            formula += "),1,0)"

            return formula

        # 这一段是用来判断是否有空单元格的，原理是遍历每一列，检查每一列的非空单元格数，若所有列的最大值和最小值相同，则认为所有信息填写完成
        def check_blank():
            # 初始化公式的基础部分
            formula = "=IF(MAX("

            # 初始化计数器，用于列名的生成
            column_letter = 'A'

            # 这里的四个范围，0，2代表A-Z，取值应该是(1,27)；1，3代表AA-A?，取值根据实际修改
            for i in range(1, 27):  # BS是Excel中的第702列
                # 生成列名（A-Z）
                column_name = chr(ord(column_letter) + i - 1)
                # 拼接COUNTA函数，注意添加范围（例如A1:A500）
                if i == 1:
                    formula += f"COUNTBLANK(数据表!{column_name}:{column_name})"
                else:
                    formula += f",COUNTBLANK(数据表!{column_name}:{column_name})"

            for i in range(1, 27):  # BS是Excel中的第702列
                # 生成列名（AA-AZ）
                column_name = chr(ord(column_letter) + i - 1)

                formula += f",COUNTBLANK(数据表!A{column_name}:A{column_name})"

            for i in range(1, 18):  # BS是Excel中的第702列
                # 生成列名（BA-BZ）
                column_name = chr(ord(column_letter) + i - 1)

                formula += f",COUNTBLANK(数据表!B{column_name}:B{column_name})"

            formula += ")=MIN("

            # 循环遍历从A到BS的所有列
            for i in range(1, 27):  # BS是Excel中的第702列
                # 生成列名（A-Z）
                column_name = chr(ord(column_letter) + i - 1)
                # 拼接COUNTA函数，注意添加范围（例如A1:A500）
                if i == 1:
                    formula += f"COUNTBLANK(数据表!{column_name}:{column_name})"
                else:
                    formula += f",COUNTBLANK(数据表!{column_name}:{column_name})"

            for i in range(1, 27):  # BS是Excel中的第702列
                # 生成列名（AA-AZ）
                column_name = chr(ord(column_letter) + i - 1)

                formula += f",COUNTBLANK(数据表!A{column_name}:A{column_name})"

            for i in range(1, 20):  # BS是Excel中的第702列
                # 生成列名（AA-AZ）
                column_name = chr(ord(column_letter) + i - 1)

                formula += f",COUNTBLANK(数据表!B{column_name}:B{column_name})"

            formula += "),1,0)"

            return formula

        # 为检查的字体添加红色
        font = Font(color="FF0000")

        # 首先判断前三列学校信息会不会有不同的
        # 列表代表了对应的列标
        for i, cell in enumerate(["A", "B", "C", "BJ"]):
            ws_check[f'C{3 * i + 1}'] = f'{cell}列是否相同'
            ws_check[f'C{3 * i + 2}'] = count_same_info(location=f"{cell}")
            ws_check[f'C{3 * i + 2}'].font = font

            cell_list.append(f'C{3 * i + 2}')

        # 然后判断下拉列表的内容是否被更改
        for i, cell in enumerate(
                ["B", "F", "G", "I", "J", "L", "M", "N", "R", "S", "T", "AB", "AD", "AE", "AG", "AK", "AM", "AO", "AP",
                 "AR", "AX", "BJ", "BK", "BL", "BM", "BN", "BO", "BR", "BS"]):
            # [item.split("_")[-1].upper() for item in lst.keys()]
            ws_check[f'E{3 * i + 1}'] = f'{cell}列下拉列表'
            ws_check[f'E{3 * i + 2}'] = gen_list_check_sentence(location=f"{cell}")
            ws_check[f'E{3 * i + 2}'].font = font

            cell_list.append(f'E{3 * i + 2}')

        # 这里判断年月格式
        for i, cell in enumerate(
                ["K", "P", "V", "X", "Y", "Z", "AA", "AC", "AI", "AJ", "AL", "AN", "AT", "AY", "AZ", "BB", "BC"]):
            ws_check[f'G{3 * i + 1}'] = f'{cell}列日期格式'
            ws_check[f'G{3 * i + 2}'] = check_str_date(location=f"{cell}")
            ws_check[f'G{3 * i + 2}'].font = font

            cell_list.append(f'G{3 * i + 2}')

        # 这里判断身份证位数
        ws_check['I1'] = 'E列身份证号位数'
        ws_check['I2'] = check_data_length(location='E', length=18)
        ws_check['I2'].font = font
        cell_list.append(f'I2')

        ws_check['I4'] = 'E列身份证号年份范围（1931-2020）'
        ws_check[
            'I5'] = '=IF(SUMPRODUCT((数据表!E2:E1000<>"")*(IFERROR(MID(数据表!E2:E1000,7,4)*1,0)>1930))=SUMPRODUCT((数据表!E2:E1000<>"")*(IFERROR(MID(数据表!E2:E1000,7,4)*1,0)<2020)),1,0)'
        ws_check['I5'].font = font
        cell_list.append(f'I5')

        ws_check['I7'] = 'E列身份证号月份范围（1-12）'
        ws_check[
            'I8'] = '=IF(SUMPRODUCT((数据表!E2:E1000<>"")*(IFERROR(MID(数据表!E2:E1000,11,2)*1,0)>0))=SUMPRODUCT((数据表!E2:E1000<>"")*(IFERROR(MID(数据表!E2:E1000,11,2)*1,0)<13)),1,0)'
        ws_check['I8'].font = font
        cell_list.append(f'I8')

        ws_check['I10'] = 'E列身份证号日期范围（1-31）'
        ws_check[
            'I11'] = '=IF(SUMPRODUCT((数据表!E2:E1000<>"")*(IFERROR(MID(数据表!E2:E1000,13,2)*1,0)>0))=SUMPRODUCT((数据表!E2:E1000<>"")*(IFERROR(MID(数据表!E2:E1000,13,2)*1,0)<32)),1,0)'
        ws_check['I11'].font = font
        cell_list.append(f'I11')

        # 手机号位数
        for i, cell in enumerate(["BE", "BI"], start=4):
            ws_check[f'I{3 * i + 1}'] = f'{cell}列手机号位数'
            ws_check[f'I{3 * i + 2}'] = check_data_length(location=f"{cell}", length=11)
            ws_check[f'I{3 * i + 2}'].font = font

            cell_list.append(f'i{3 * i + 2}')

        # LM列参加工作前学历学位对照
        ws_check[f'K1'] = "L/M列参加工作前学历学位-无学位学历"
        ws_check[
            f'K2'] = '=IF(SUMPRODUCT(ISNUMBER(MATCH(数据表!L:L,{"初中","高中","中师","中专（非师范）","专科"},0))*ISNUMBER(MATCH(数据表!M:M,{"无"},0)))=SUM(COUNTIF(数据表!L:L,{"初中","高中","中师","中专（非师范）","专科"})),1,0)'
        ws_check[f'K2'].font = font
        cell_list.append(f'K2')

        ws_check[f'K4'] = "L/M列参加工作前学历学位-本科"
        ws_check[
            f'K5'] = '=IF(SUMPRODUCT(ISNUMBER(MATCH(数据表!L:L,{"本科"},0))*ISNUMBER(MATCH(数据表!M:M,{"学士学位","无"},0)))=SUM(COUNTIF(数据表!L:L,{"本科"})),1,0)'
        ws_check[f'K5'].font = font
        cell_list.append(f'K5')

        ws_check[f'K7'] = "L/M列参加工作前学历学位-硕士研究生"
        ws_check[
            f'K8'] = '=IF(SUMPRODUCT(ISNUMBER(MATCH(数据表!L:L,{"硕士研究生"},0))*ISNUMBER(MATCH(数据表!M:M,{"硕士学位","无"},0)))=SUM(COUNTIF(数据表!L:L,{"硕士研究生"})),1,0)'
        ws_check[f'K8'].font = font
        cell_list.append(f'K8')

        ws_check[f'K10'] = "L/M列参加工作前学历学位-博士研究生"
        ws_check[
            f'K11'] = '=IF(SUMPRODUCT(ISNUMBER(MATCH(数据表!L:L,{"博士研究生"},0))*ISNUMBER(MATCH(数据表!M:M,{"博士学位","无"},0)))=SUM(COUNTIF(数据表!L:L,{"博士研究生"})),1,0)'
        ws_check[f'K11'].font = font
        cell_list.append(f'K11')

        # RS列最高学历对应学位对照
        ws_check[f'K13'] = "R/S列最高学历对应学位-无学位学历"
        ws_check[
            f'K14'] = '=IF(SUMPRODUCT(ISNUMBER(MATCH(数据表!R:R,{"初中","高中","中师","中专（非师范）","专科"},0))*ISNUMBER(MATCH(数据表!S:S,{"无"},0)))=SUM(COUNTIF(数据表!R:R,{"初中","高中","中师","中专（非师范）","专科"})),1,0)'
        ws_check[f'K14'].font = font
        cell_list.append(f'K14')

        ws_check[f'K16'] = "R/S列最高学历对应学位-本科"
        ws_check[
            f'K17'] = '=IF(SUMPRODUCT(ISNUMBER(MATCH(数据表!R:R,{"本科"},0))*ISNUMBER(MATCH(数据表!S:S,{"学士学位","无"},0)))=SUM(COUNTIF(数据表!R:R,{"本科"})),1,0)'
        ws_check[f'K17'].font = font
        cell_list.append(f'K17')

        ws_check[f'K19'] = "R/S列最高学历对应学位-硕士研究生"
        ws_check[
            f'K20'] = '=IF(SUMPRODUCT(ISNUMBER(MATCH(数据表!R:R,{"硕士研究生"},0))*ISNUMBER(MATCH(数据表!S:S,{"硕士学位","无"},0)))=SUM(COUNTIF(数据表!R:R,{"硕士研究生"})),1,0)'
        ws_check[f'K20'].font = font
        cell_list.append(f'K20')

        ws_check[f'K22'] = "R/S列最高学历对应学位-博士研究生"
        ws_check[
            f'K23'] = '=IF(SUMPRODUCT(ISNUMBER(MATCH(数据表!R:R,{"博士研究生"},0))*ISNUMBER(MATCH(数据表!S:S,{"博士学位","无"},0)))=SUM(COUNTIF(数据表!R:R,{"博士研究生"})),1,0)'
        ws_check[f'K23'].font = font
        cell_list.append(f'K23')

        # 填写完成的判断
        ws_check['A1'] = '检查无误'
        ws_check['A2'] = f'=IF(AND({"=1,".join(["A5", "A8"] + cell_list)}=1),1,0)'
        ws_check['A2'].font = font

        ws_check['A4'] = '是否已填写完成'
        ws_check['A5'] = check_blank()
        ws_check['A5'].font = font

        ws_check['A7'] = '公式是否完整'
        ws_check['A8'] = check_formula(
            td_list=cell_list)
        ws_check['A8'].font = font

    # 这里是函数第一句
    wb = Workbook()
    ws = wb.active
    ws.title = "数据表"

    ws_check = wb.create_sheet("公式表")
    ws_list = wb.create_sheet("下拉列表信息")

    ws.append(title)

    area = area_name

    name_list = []
    for data_row in data:
        name_list.append(data_row[3])

    for data_row in data:
        ws.append(list(map(str, data_row)))

    add_restrict()

    add_bg_colour(ws=ws)

    add_excel_func()

    set_col_width(ws=ws)
    set_col_width(ws=ws_check)
    set_col_width(ws=ws_list)

    add_sheet_limit(sheet=ws_check)
    add_sheet_limit(sheet=ws_list)

    set_all_cell_borders(ws)

    # 将整个工作表的默认数字格式设置为文本
    for row in ws.iter_rows():
        for cell in row:
            cell.number_format = numbers.FORMAT_TEXT

    ensure_folders_exist_or_clear(kind="在编", school_name1=file_name, area_name1=area_name)

    write_list_to_txt(
        route=fr"C:\Users\1012986131\Desktop\python\streamlit_pyecharts\update_data_source\2025学年\output\在编\{area}\{file_name}\下拉列表内容.txt")

    write_name_list_to_txt(
        route=fr"C:\Users\1012986131\Desktop\python\streamlit_pyecharts\update_data_source\2025学年\output\在编\{area}\{file_name}\已有人员名单.txt",
        name_list=name_list)

    wb.save(
        fr"C:\Users\1012986131\Desktop\python\streamlit_pyecharts\update_data_source\2025学年\output\在编\{area}\{file_name}\{file_name}.xlsx")


if __name__ == '__main__':
    data = read_xlsx_to_list(
        file_path=fr"C:\Users\1012986131\Desktop\python\streamlit_pyecharts\update_data_source\2025学年\data\teacher_0.xlsx")[
        1:]

    output = {item: {} for item in ["直管", "新市", "永平", "石井", "江高", "人和", "太和", "钟落潭"]}

    for item in data:
        item[-1] = ""
        item.extend(["", "", ""])
        if item[0] not in output[item[-11]].keys():
            output[item[-11]][item[0]] = []

        output[item[-11]][item[0]].append(item)

    with open(fr"test.json", "w", encoding="UTF-8") as f:
        # 将生成的数据保存至json文件中
        json.dump(output, f, indent=4, ensure_ascii=False)

    for area, temp in output.items():
        for school, data_temp in temp.items():
            print(f"正在处理/{area}/{school}/的数据")
            output_excel_0(title=title_0, data=data_temp, file_name=school, area_name=area)
