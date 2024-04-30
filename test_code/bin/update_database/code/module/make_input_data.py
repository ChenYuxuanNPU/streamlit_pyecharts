#
# 这个文件获取表格中的数据并验证长度
# 不需要单独跑，被data_insert调用
#

import openpyxl


def read_input_data(kind):
    # 确定收集的数据表
    if kind == "在编":
        wb = openpyxl.load_workbook('C:\\Users\\1012986131\\Desktop\\python\\streamlit_pyecharts\\update_database'
                                    '\\data_source\\data_0.xlsx', data_only=True)
        sheet = wb["Sheet1"]

    elif kind == "非编":
        wb = openpyxl.load_workbook('C:\\Users\\1012986131\\Desktop\\python\\streamlit_pyecharts\\update_database'
                                    '\\data_source\\data_1.xlsx', data_only=True)
        sheet = wb["data0"]

    else:
        print(r"kind参数错误(make_input_data.py)")
        return -1

    # 获取表格数据
    result = []
    for row in sheet.iter_rows(values_only=True):
        # 不读取第一行列标
        if not str(row[0]).isnumeric():
            result.append(row)
            # print(row)

    # 用来验证所有数据是否等长
    flag = 0
    for i in range(1, len(result)):
        if len(result[i]) != len(result[0]):
            flag = 1

    if flag == 0:
        print(fr"单条数据长度：{str(len(result[0]))}（make_input_data.py）")
        print(fr"总数据量：{str(len(result))}（make_input_data.py）")

        return result

    else:
        print(r"数据长度不相同（make_input_data.py）")

        return -2
