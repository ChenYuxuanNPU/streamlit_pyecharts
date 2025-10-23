import re
import sqlite3
from datetime import datetime

import openpyxl
from dateutil.relativedelta import relativedelta


def del_tuple_in_list(data: list) -> list:
    """
    将形如[('1',), ('2',), ('3',),]的数据转化为[1, 2, 3,]
    :param data:带有元组的列表
    :return: 清洗后的列表
    """

    output = []

    output.extend(single_data[0] for single_data in data)

    return output


def execute_sql_sentence(sentence: str, ) -> list:
    """
    执行数据库语句并返回列表
    :param sentence: 需要执行的语句
    :return:
    """

    c, conn = connect_database()

    result = []
    try:
        c.execute(sentence)
        result = c.fetchall()
        conn.commit()

    except Exception as e:
        print(e)

    disconnect_database(conn=conn)

    return result


# kind:"在编","编外"
def connect_database() -> tuple[sqlite3.Cursor, sqlite3.Connection]:
    """
    用于连接数据库
    :return:
    """

    conn = sqlite3.connect(
        fr"C:\Users\1012986131\Desktop\python\streamlit_pyecharts\database\educational_data.db"
    )
    c = conn.cursor()

    return c, conn


def disconnect_database(conn) -> None:
    """
    用于断开数据库
    :param conn:
    :return:
    """

    conn.close()

    return None


def remove_zero_width_chars(text):
    # 移除所有零宽字符（包括\u200c, \u200d, \u200e, \u200f, \uFEFF等）
    return re.sub(r'[\u200c\u200d\u200e\u200f\uFEFF]', '', text)


def read_xlsx_to_list(file_path, sheet_name=None):
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active if sheet_name is None else workbook[sheet_name]

    data = []
    for row in sheet.iter_rows(values_only=True):
        processed_row = []
        for cell_value in row:
            if isinstance(cell_value, str):
                cell_value = remove_zero_width_chars(cell_value).strip()
            processed_row.append(cell_value)
        data.append(processed_row)

    return data


def save_excel(two_dimension_list: list[list or tuple], excel_name: str = "output"):
    # 读取 Excel 文件
    workbook = openpyxl.Workbook()

    # 获取活动的工作表
    ws = workbook.active

    for item in two_dimension_list:
        ws.append(item)

    workbook.save(f"{excel_name}.xlsx")


def get_age_from_citizen_id(citizen_id: str, year: str = None, month: int = 9, day: int = 1) -> int:
    """
    通过身份证号计算当前年龄或截止某一年某一月某一日（默认某一年的9月1日）\n
    get_age_from_citizen_id(citizen_id = "440105200102220000") -> 23
    :param citizen_id: 身份证号
    :param year: 截止年份
    :param month: 截止月份
    :param day: 截止日期
    :return: 年龄,两位数int
    """

    if len(citizen_id) != 18:
        return -1

    try:
        if year is None:
            return max(
                relativedelta(
                    dt1=datetime.today(),
                    dt2=datetime(
                        year=int(citizen_id[6:10]),
                        month=int(citizen_id[10:12]),
                        day=int(citizen_id[12:14])
                    )
                ).years,
                0
            )

        elif 2000 <= int(year) <= 3000:
            return max(
                relativedelta(
                    dt1=datetime(year=int(year), month=month, day=day),
                    dt2=datetime(
                        year=int(citizen_id[6:10]),
                        month=int(citizen_id[10:12]),
                        day=int(citizen_id[12:14])
                    )

                ).years,
                0
            )

        else:
            return -2

    except Exception as e:
        print(f"{e}:{citizen_id}")
        return -3


def get_educational_background_order() -> dict:
    """
    学历排序
    :return:
    """

    return {
        '博士研究生': 1, '硕士研究生': 2, '本科': 3, "专科": 4, "高中": 5, "高中及以下": 6,
        "中师": 7, "中专（非师范）": 8, "中专": 9, "初中": 10, None: 11
    }


def get_period_order() -> dict:
    """
    学段排序
    :return:
    """

    return {'高中': 1, '初中': 2, '小学': 3, '幼儿园': 4, '中职': 5, "其他": 6, None: 7, }


def get_code_of_985() -> list[str]:
    """
    985院校代码列表
    :return:
    """
    return ['10003', '10001', '10614', '10335', '10384', '10533', '10558', '10486', '10246', '10487', '10284', '10286',
            '10610', '10247', '10055', '10422', '10002', '10248', '10561', '10183', '10269', '10532', '10611', '10698',
            '10213', '18213', '10358', '10423', '10141', '10056', '10027', '10145', '10007', '10006', '10730', '10699',
            '10712', '10019', '10052', '19248', '91002', '19246', '7321']


def get_code_of_211() -> list[str]:
    """
    211院校代码列表
    :return:
    """
    return ['10003', '10001', '10614', '10335', '10384', '10533', '10558', '10486', '10246', '10487', '10284', '10286',
            '10610', '10247', '10055', '10422', '10002', '10248', '10561', '10183', '10269', '10532', '10611', '10698',
            '10213', '18213', '10358', '10423', '10141', '10056', '10027', '10145', '10007', '10006', '10730', '10699',
            '10712', '10019', '10052', '19248', '91002', '19246', '7321', '10635', '10559', '10033', '10280',
            '10285', '10613', '10497', '10459', '10295', '10520', '10697', '10255', '10403', '10651', '10294', '10290',
            '10030', '10511', '10589', '10251', '10359', '19359', '10710', '10701', '10288', '10272', '10054', '10079',
            '10008', '10287', '10004', '10386', '10053', '10574', '10036', '10034', '10319', '10357', '10080', '10718',
            '10217', '10013', '10673', '10005', '10542', '10200', '10593', '10140', '10112', '10151', '10504', '10657',
            '10271', '10626', '10022', '10759', '10184', '10010', '10045', '10307', '10316', '10225', '10043', '10126',
            '10026', '10755', '10224', '10749', '10425', '10062', '10694', '10743', '11414', '10491', '11413', '11415',
            '19635', '19414', '91030', '90026']


def get_sheet_names(file_path: str) -> list:
    wb = openpyxl.load_workbook(file_path, read_only=True)
    names = wb.sheetnames
    wb.close()
    return names


def generate_subsets(arr):
    n = len(arr)
    res = []
    # 遍历所有非空子集（从1到2^n - 1）
    for i in range(1, 1 << n):
        subset = []
        for j in range(n):
            if i & (1 << j):
                subset.append(arr[j])
        # 将子集列表转换为逗号分隔的字符串
        res.append('，'.join(subset))
    return res


def generate_letter_sequence(start, end):
    """
    生成从起始字母到结束字母的所有字母序列
    :param start: 起始字母(长度为1-3的大写字母)
    :param end: 结束字母(长度为1-3的大写字母)
    :return: 按顺序排列的所有字母组合
    """

    # 将字母转换为数字表示(A=1, B=2, ..., Z=26)
    def letter_to_number(letter):
        return sum((ord(char) - 64) * (26 ** i) for i, char in enumerate(reversed(letter)))

    # 将数字转换回字母
    def number_to_letter(num):
        result = []
        while num > 0:
            num -= 1
            result.append(chr((num % 26) + 65))
            num //= 26
        return ''.join(reversed(result)) if result else 'A'

    start_num = letter_to_number(start)
    end_num = letter_to_number(end)

    # 生成序列
    sequence = []
    for num in range(start_num, end_num + 1):
        sequence.append(number_to_letter(num))

    return sequence
