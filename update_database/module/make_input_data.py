#
# 这个文件获取表格中的数据并验证长度
# 不需要单独跑，被data_insert调用
#
import json
from pathlib import Path

import openpyxl


def get_database_basic_info() -> dict:
    """
    读取数据库基本信息
    :return: 数据库信息字典
    """

    with open(fr"{Path(__file__).resolve().parent.parent.parent}\json_file\database\database_basic_info.json", "r", encoding="UTF-8") as f:
        json_data = json.load(f)

    return json_data


def read_input_data(kind) -> list:
    """
    用于从数据源的表格中提取数据
    :param kind: 在编/编外
    :return: 二维列表，每一行代表一个老师的信息
    """

    result = []
    json_data = get_database_basic_info()

    # 字典里0位是数据源文件名，1位是表名
    wb = openpyxl.load_workbook(fr'{Path(__file__).resolve().parent.parent.parent}\update_database'
                                fr'\data_source\{json_data["xlsx_file_and_sheet_name"][kind][0]}', data_only=True)
    sheet = wb[f"{json_data["xlsx_file_and_sheet_name"][kind][1]}"]

    # 确定收集的数据表
    if "教师信息" in kind:

        # 获取表格数据
        for row in sheet.iter_rows(values_only=True):
            # 不读取第一行列标
            if not str(row[0]).isnumeric():
                result.append(row)
                # print(row)

    elif "学校信息总览" in kind:

        # 获取表格数据
        for row in sheet.iter_rows(values_only=True):

            # 不读取第一行列标
            if str(row[0]).isnumeric():
                result.append(row)

    else:
        print(fr"kind参数错误:{kind} (make_input_data.py)")
        return [-1]

    # 判断一下result是否为空值
    if len(result) == 0:
        print("读取的数据为空值 (make_input_data.py)")
        return [-1]

    # 用来验证所有数据是否等长

    if max(len(item) for item in result) == min(len(item) for item in result):

        print(fr"{kind}单条数据长度:{str(len(result[0]))} (make_input_data.py)")
        print(fr"{kind}总数据量:{str(len(result))} (make_input_data.py)")

        return result

    else:
        print(fr"{kind}数据长度不相同 (make_input_data.py)")

        return [-1]


if __name__ == '__main__':
    print(read_input_data(kind="2023年学校信息总览"))
    # print(get_nth_parent_dir(n=3))
